from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = PROJECT_ROOT / "docs" / "codex-brain" / "magok_ksic11_full_directory.csv"
XLSX_PATH = PROJECT_ROOT / "docs" / "codex-brain" / "magok_ksic11_full_directory.xlsx"


def main() -> None:
    with CSV_PATH.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.reader(csv_file)
        rows = list(reader)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Magok KSIC 11"

    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="DDEBFF")
    header_font = Font(bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_length + 2, 12), 42)

    workbook.save(XLSX_PATH)


if __name__ == "__main__":
    main()
